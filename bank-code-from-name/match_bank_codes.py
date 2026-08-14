"""根据收款开户行名称匹配支付系统行号。

收款开户行常会使用简称，例如“工行北京新华支行”。本模块先按名称逐字
缩小主数据工作表范围，再在候选工作表的 ``BnkName|银行名称`` 中逐字
缩小候选记录范围，最后仅在结果可唯一确定时写入支付系统行号。
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from copy import copy
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


MASTER_CODE_HEADERS = ("PaySysBnkCode|支付系统行号", "PaySysBnkCode")
MASTER_NAME_HEADERS = ("BnkName|银行名称", "BnkName")
TARGET_NAME_HEADERS = (
    "收款开户行",
    "原收款行名称",
    "收款行名称",
    "开户银行",
    "开户行",
    "收款银行",
    "银行名称",
    "BnkName|银行名称",
    "BnkName",
)
DEFAULT_OUTPUT_COLUMN = "PaySysBnkCode|支付系统行号"
DEFAULT_MATCHED_NAME_COLUMN = "匹配收款开户行"
DEFAULT_STATUS_COLUMN = "匹配状态"
DEFAULT_CANDIDATE_COLUMN = "匹配候选结果"
PROJECT_DIRECTORY = Path(__file__).resolve().parent.parent

ProgressCallback = Callable[[int, int, dict[str, int]], None]
CancelCallback = Callable[[], bool]


# 这些是供应商资料中最常见的银行简称。未列出的名称仍会走逐字匹配。
BANK_ABBREVIATIONS = {
    "工行": "中国工商银行",
    "农行": "中国农业银行",
    "中行": "中国银行",
    "建行": "中国建设银行",
    "交行": "交通银行",
    "邮储": "中国邮政储蓄银行",
    "邮政储蓄": "中国邮政储蓄银行",
    "招行": "招商银行",
    "浦发": "上海浦东发展银行",
    "民生": "中国民生银行",
    "光大": "中国光大银行",
    "兴业": "兴业银行",
    "平安": "平安银行",
    "中信": "中信银行",
    "华夏": "华夏银行",
    "广发": "广发银行",
    "浙商": "浙商银行",
    "渤海": "渤海银行",
    "恒丰": "恒丰银行",
}

class MatchingCancelled(Exception):
    """用户在匹配完成前取消了任务。"""


@dataclass(frozen=True)
class BankRecord:
    code: str
    name: str
    match_name: str
    reduced_name: str


@dataclass
class MasterSheet:
    title: str
    match_title: str
    records: list[BankRecord]
    record_ids_by_reduced_name: dict[str, list[int]]


@dataclass(frozen=True)
class MatchResult:
    code: str | None
    status: str
    candidate_count: int = 0
    candidates: tuple[tuple[str, str], ...] = ()
    matched_name: str | None = None


def normalize_code(value: object) -> str:
    """将单元格中的行号统一为 12 位数字文本，保留前导零。"""
    if value is None:
        return ""

    text = str(value).strip()
    if not text:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text.zfill(12) if text.isdigit() and len(text) < 12 else text


def header_positions(sheet: Worksheet) -> dict[str, int]:
    """返回首行非空表头及其列号（从 1 开始）。"""
    return {
        str(cell.value).strip(): cell.column
        for row in sheet.iter_rows(min_row=1, max_row=1)
        for cell in row
        if cell.value is not None and str(cell.value).strip()
    }


def find_column(headers: dict[str, int], candidates: Iterable[str]) -> int | None:
    """按候选列名顺序返回第一个存在的列号。"""
    for candidate in candidates:
        if candidate in headers:
            return headers[candidate]
    return None


def sheet_prefix(sheet: Worksheet) -> str:
    """工作表名称的前三位数字，例如“001中国人民银行”返回“001”。"""
    title_prefix = sheet.title[:3]
    return title_prefix if title_prefix.isdigit() else ""


def normalize_bank_name(value: object) -> str:
    """规范化银行名称，并将常用简称展开为全称。"""
    if value is None:
        return ""

    text = str(value).strip()
    if not text:
        return ""
    text = re.sub(r"[\s\-_\/\\()（）\[\]【】{}<>《》,，.。;；:：'\"`~!！?？]+", "", text)
    for abbreviation, full_name in BANK_ABBREVIATIONS.items():
        # 已经是“中国民生银行”这类全称时，不应再把“民生”替换一次。
        text = re.sub(
            rf"(?<!中国){re.escape(abbreviation)}(?!银行)", full_name, text
        )
    return text


def reduce_legal_suffixes(value: str) -> str:
    """移除公司法律形式，便于比较简称与主数据全称。"""
    for suffix in ("股份有限公司", "有限责任公司", "有限公司"):
        value = value.replace(suffix, "")
    return value


def copy_header_style(source: Cell, destination: Cell) -> None:
    """新增列沿用输入列的表头样式。"""
    if source.has_style:
        destination._style = copy(source._style)
    if source.number_format:
        destination.number_format = source.number_format
    if source.font:
        destination.font = copy(source.font)
    if source.fill:
        destination.fill = copy(source.fill)
    if source.border:
        destination.border = copy(source.border)
    if source.alignment:
        destination.alignment = copy(source.alignment)
    if source.protection:
        destination.protection = copy(source.protection)


def get_or_add_column(sheet: Worksheet, header: str, style_source: Cell) -> int:
    """返回表头列；不存在时在末列右侧追加。"""
    headers = header_positions(sheet)
    if header in headers:
        return headers[header]

    column = sheet.max_column + 1
    cell = sheet.cell(row=1, column=column, value=header)
    copy_header_style(style_source, cell)
    return column


def _load_legacy_xls(path: Path) -> Workbook:
    """将旧版 xls 读取为可写入结果的 xlsx 工作簿。

    xls 无法由 openpyxl 直接编辑，因此读取原始单元格值、列宽和行高后，
    以 xlsx 工作簿承载匹配结果。结果文件必须使用 .xlsx 扩展名。
    """
    try:
        import xlrd
    except ImportError as error:
        raise ValueError(
            "读取 .xls 文件需要 xlrd。请执行：python -m pip install -r requirements.txt"
        ) from error

    source = xlrd.open_workbook(path, formatting_info=False)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for source_sheet in source.sheets():
        sheet = workbook.create_sheet(source_sheet.name)
        for row_index in range(source_sheet.nrows):
            for column_index in range(source_sheet.ncols):
                cell = source_sheet.cell(row_index, column_index)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(value, source.datemode)
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    value = xlrd.error_text_from_code.get(value, "#VALUE!")
                sheet.cell(row=row_index + 1, column=column_index + 1, value=value)

        for column_index, column_info in source_sheet.colinfo_map.items():
            sheet.column_dimensions[get_column_letter(column_index + 1)].width = (
                column_info.width / 256
            )
        for row_index, row_info in source_sheet.rowinfo_map.items():
            if row_info.height:
                sheet.row_dimensions[row_index + 1].height = row_info.height / 20
    return workbook


def load_target_workbook(target_path: Path) -> Workbook:
    """载入 xlsx/xlsm 或旧版 xls 待匹配文件。"""
    if target_path.suffix.lower() == ".xls":
        return _load_legacy_xls(target_path)
    return load_workbook(target_path)


def read_master_data(master_path: Path) -> list[MasterSheet]:
    """读取行名行号主数据并建立工作表及逐字索引。"""
    if master_path.suffix.lower() == ".xls":
        raise ValueError("行名行号主数据请使用 .xlsx 或 .xlsm 文件。")

    workbook = load_workbook(master_path, read_only=True, data_only=True)
    master_sheets: list[MasterSheet] = []
    try:
        for sheet in workbook.worksheets:
            headers = header_positions(sheet)
            code_column = find_column(headers, MASTER_CODE_HEADERS)
            name_column = find_column(headers, MASTER_NAME_HEADERS)
            if code_column is None or name_column is None:
                continue

            records: list[BankRecord] = []
            reduced_name_index: dict[str, list[int]] = defaultdict(list)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                code = normalize_code(row[code_column - 1])
                name_value = row[name_column - 1]
                name = "" if name_value is None else str(name_value).strip()
                match_name = normalize_bank_name(name)
                if not (len(code) == 12 and code.isdigit() and match_name):
                    continue

                record_id = len(records)
                reduced_name = reduce_legal_suffixes(match_name)
                records.append(
                    BankRecord(
                        code=code,
                        name=name,
                        match_name=match_name,
                        reduced_name=reduced_name,
                    )
                )
                reduced_name_index[reduced_name].append(record_id)

            if records:
                master_sheets.append(
                    MasterSheet(
                        title=sheet.title,
                        match_title=normalize_bank_name(sheet.title[3:]),
                        records=records,
                        record_ids_by_reduced_name=dict(reduced_name_index),
                    )
                )
    finally:
        workbook.close()

    if not master_sheets:
        raise ValueError(
            "未在主数据文件中找到 PaySysBnkCode|支付系统行号 和 BnkName|银行名称 列。"
        )
    return master_sheets


def _find_candidate_sheets(
    bank_name: str, master_sheets: list[MasterSheet]
) -> list[MasterSheet]:
    """先精确定位银行主体 Sheet；未命中时才按简称逐字收缩候选范围。"""
    first_bank_end = bank_name.find("行")
    bank_subject = bank_name if first_bank_end < 0 else bank_name[: first_bank_end + 1]
    exact_sheets = [
        sheet for sheet in master_sheets if sheet.match_title == bank_subject
    ]
    if exact_sheets:
        return exact_sheets

    candidates = master_sheets
    for character_index, character in enumerate(bank_name):
        narrowed = [sheet for sheet in candidates if character in sheet.match_title]
        if not narrowed:
            if character_index == 0:
                return []
            break
        candidates = narrowed
        if len(candidates) == 1:
            break
    return candidates


def _candidate_records_in_order(
    branch_name: str, sheets: list[MasterSheet]
) -> list[BankRecord]:
    """按顺序为每个候选累积网点名称匹配分，保留最高分候选。"""
    candidates = [
        (
            record,
            record.reduced_name.find("行"),
            record.reduced_name.find("行") + 1,
            0,
        )
        for sheet in sheets
        for record in sheet.records
    ]
    for character_index, character in enumerate(branch_name):
        weight = len(branch_name) - character_index
        updated: list[tuple[BankRecord, int, int, int]] = []
        for record, position, branch_start, score in candidates:
            matched_position = record.reduced_name.find(character, position + 1)
            if matched_position >= 0:
                updated.append((record, matched_position, branch_start, score + weight))
            else:
                updated.append((record, position, branch_start, score))
        candidates = updated

    best_score = max((score for _, _, _, score in candidates), default=0)
    if best_score == 0:
        return []
    return [record for record, _, _, score in candidates if score == best_score]


def _candidate_records(
    bank_name: str, sheets: list[MasterSheet]
) -> list[BankRecord]:
    """仅按网点名称的正序字符筛选候选记录。"""
    bank_subject_end = bank_name.find("行") + 1
    branch_name = bank_name[bank_subject_end:]
    return _candidate_records_in_order(branch_name, sheets)


def _direct_records_in_sheets(
    sheets: list[MasterSheet], reduced_name: str
) -> list[BankRecord]:
    """在已确定的银行 Sheet 中查找完整名称精确匹配的记录。"""
    return [
        sheet.records[record_id]
        for sheet in sheets
        for record_id in sheet.record_ids_by_reduced_name.get(reduced_name, ())
    ]


def _ordered_detail_match_count(
    bank_name: str, record: BankRecord
) -> tuple[int, int]:
    """Return ordered matches in the branch portion of a bank name."""
    bank_subject_end = bank_name.find("行") + 1
    detail_length = len(bank_name) - bank_subject_end
    if detail_length <= 0:
        return 0, 0

    branch_name = bank_name[bank_subject_end:]
    branch_start = record.reduced_name.find("行") + 1
    position = branch_start - 1
    matched_detail_characters = 0
    for character in branch_name:
        matched_position = record.reduced_name.find(character, position + 1)

        if matched_position >= 0:
            position = matched_position
            matched_detail_characters += 1

    return matched_detail_characters, detail_length


def _has_sufficient_ordered_detail(
    bank_name: str, records: list[BankRecord]
) -> bool:
    """Require 75 percent ordered coverage after the bank subject for auto-match."""
    for record in records:
        matched, total = _ordered_detail_match_count(bank_name, record)
        if total and matched * 4 >= total * 3:
            return True
    return False


def _ambiguous_result(records: list[BankRecord]) -> MatchResult:
    """返回去重后的候选银行名称与支付系统行号，供人工确认。"""
    candidates_by_code: dict[str, str] = {}
    for record in records:
        candidates_by_code.setdefault(record.code, record.name)
    candidates = tuple(
        (name, code) for code, name in sorted(candidates_by_code.items())
    )
    return MatchResult(
        None,
        f"匹配不唯一（{len(candidates)} 个候选）",
        len(candidates),
        candidates,
    )


def _low_confidence_result(records: list[BankRecord]) -> MatchResult:
    """Expose a weak similarity candidate without using its payment code."""
    ambiguous = _ambiguous_result(records)
    return MatchResult(
        None,
        f"匹配置信度不足（{ambiguous.candidate_count} 个候选）",
        ambiguous.candidate_count,
        ambiguous.candidates,
    )


def format_candidate_results(candidates: tuple[tuple[str, str], ...]) -> str | None:
    """将候选项格式化为“银行名称：支付系统行号”，避免超过 Excel 单元格上限。"""
    if not candidates:
        return None

    parts: list[str] = []
    total_length = 0
    for index, (name, code) in enumerate(candidates):
        item = f"{name}：{code}"
        separator_length = 1 if parts else 0
        if total_length + separator_length + len(item) > 32_000:
            parts.append(f"……其余 {len(candidates) - index} 个候选未显示")
            break
        parts.append(item)
        total_length += separator_length + len(item)
    return "\n".join(parts)


def match_bank_name(bank_name_value: object, master_sheets: list[MasterSheet]) -> MatchResult:
    """依据收款开户行名称返回可安全写入的支付系统行号。"""
    bank_name = normalize_bank_name(bank_name_value)
    if not bank_name:
        return MatchResult(None, "收款开户行为空")

    reduced_name = reduce_legal_suffixes(bank_name)
    candidate_sheets = _find_candidate_sheets(bank_name, master_sheets)
    if not candidate_sheets:
        return MatchResult(None, "未找到")

    # A complete name is decisive even when a longer branch shares the same
    # prefix.  Do this before character narrowing so an unrelated extension
    # cannot hide the exact branch during forward matching.
    direct_records = _direct_records_in_sheets(candidate_sheets, reduced_name)
    direct_codes = {record.code for record in direct_records}
    if len(direct_codes) == 1:
        direct_record = next(
            record for record in direct_records if record.code in direct_codes
        )
        return MatchResult(
            direct_record.code, "已匹配", matched_name=direct_record.name
        )
    if direct_records:
        return _ambiguous_result(direct_records)

    records = _candidate_records(reduced_name, candidate_sheets)
    if not records:
        return MatchResult(None, "未找到")

    codes = {record.code for record in records}
    if len(codes) == 1:
        if not _has_sufficient_ordered_detail(reduced_name, records):
            return _low_confidence_result(records)
        matched_record = next(record for record in records if record.code in codes)
        return MatchResult(
            matched_record.code, "已匹配", matched_name=matched_record.name
        )
    return _ambiguous_result(records)


def process_target_file(
    target_path: Path,
    output_path: Path,
    master_sheets: list[MasterSheet],
    output_column: str = DEFAULT_OUTPUT_COLUMN,
    status_column: str = DEFAULT_STATUS_COLUMN,
    candidate_column: str = DEFAULT_CANDIDATE_COLUMN,
    matched_name_column: str = DEFAULT_MATCHED_NAME_COLUMN,
    *,
    progress_callback: ProgressCallback | None = None,
    cancel_callback: CancelCallback | None = None,
    progress_every: int = 1_000,
    target_sheet_names: set[str] | None = None,
) -> dict[str, int]:
    """将匹配到的支付系统行号写入结果文件，并返回统计数。"""
    if progress_every < 1:
        raise ValueError("progress_every 必须大于 0。")
    if output_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("结果文件必须使用 .xlsx 或 .xlsm 扩展名。")

    workbook = load_target_workbook(target_path)
    summary: defaultdict[str, int] = defaultdict(int)
    target_sheets: list[tuple[Worksheet, int]] = []
    total_rows = 0
    try:
        if target_sheet_names is not None:
            available_sheet_names = {sheet.title for sheet in workbook.worksheets}
            missing_sheet_names = target_sheet_names - available_sheet_names
            if missing_sheet_names:
                names = "、".join(sorted(missing_sheet_names))
                raise ValueError(f"待匹配文件中不存在指定工作表：{names}。")

        for sheet in workbook.worksheets:
            if target_sheet_names is not None and sheet.title not in target_sheet_names:
                continue
            name_column = find_column(header_positions(sheet), TARGET_NAME_HEADERS)
            if name_column is None:
                continue
            target_sheets.append((sheet, name_column))
            total_rows += max(sheet.max_row - 1, 0)

        if not target_sheets:
            accepted_headers = "、".join(TARGET_NAME_HEADERS)
            raise ValueError(f"待匹配文件中未找到收款开户行列。可识别列名：{accepted_headers}。")

        if progress_callback:
            progress_callback(0, total_rows, dict(summary))

        # 供应商表中同一开户行通常出现多次，缓存可避免重复匹配大量主数据记录。
        match_cache: dict[str, MatchResult] = {}
        processed_rows = 0
        for sheet, name_column in target_sheets:
            if cancel_callback and cancel_callback():
                raise MatchingCancelled("用户已取消匹配。")

            style_source = sheet.cell(row=1, column=name_column)
            result_column = get_or_add_column(sheet, output_column, style_source)
            matched_name_column_index = get_or_add_column(
                sheet, matched_name_column, style_source
            )
            status_column_index = get_or_add_column(sheet, status_column, style_source)
            candidate_column_index = get_or_add_column(sheet, candidate_column, style_source)

            for row_number in range(2, sheet.max_row + 1):
                bank_name = sheet.cell(row=row_number, column=name_column).value
                cache_key = normalize_bank_name(bank_name)
                result = match_cache.get(cache_key)
                if result is None:
                    result = match_bank_name(bank_name, master_sheets)
                    match_cache[cache_key] = result

                result_cell = sheet.cell(row=row_number, column=result_column)
                matched_name_cell = sheet.cell(
                    row=row_number, column=matched_name_column_index
                )
                status_cell = sheet.cell(row=row_number, column=status_column_index)
                candidate_cell = sheet.cell(row=row_number, column=candidate_column_index)
                result_cell.value = result.code
                matched_name_cell.value = result.matched_name
                candidate_cell.value = format_candidate_results(result.candidates)
                if result.code:
                    result_cell.number_format = "@"
                    summary["matched"] += 1
                elif result.status == "收款开户行为空":
                    summary["empty"] += 1
                elif result.candidate_count:
                    summary["ambiguous"] += 1
                else:
                    summary["not_found"] += 1
                status_cell.value = result.status

                processed_rows += 1
                if cancel_callback and cancel_callback():
                    raise MatchingCancelled("用户已取消匹配。")
                if progress_callback and (
                    processed_rows % progress_every == 0 or processed_rows == total_rows
                ):
                    progress_callback(processed_rows, total_rows, dict(summary))

        workbook.save(output_path)
        summary["sheets"] = len(target_sheets)
        return dict(summary)
    finally:
        workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="根据收款开户行名称匹配支付系统行号。")
    parser.add_argument(
        "master",
        nargs="?",
        type=Path,
        default=PROJECT_DIRECTORY / "行名行号2026.03.01.xlsx",
        help="银行行名行号主数据 .xlsx/.xlsm 文件。",
    )
    parser.add_argument(
        "target",
        nargs="?",
        type=Path,
        default=PROJECT_DIRECTORY / "供应商库.xls",
        help="待匹配 Excel 文件，支持 .xls、.xlsx、.xlsm。",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="结果文件路径；默认生成“原文件名_匹配结果.xlsx”。",
    )
    parser.add_argument(
        "--output-column",
        default=DEFAULT_OUTPUT_COLUMN,
        help=f"写入支付系统行号的列名，默认：{DEFAULT_OUTPUT_COLUMN}。",
    )
    parser.add_argument(
        "--status-column",
        default=DEFAULT_STATUS_COLUMN,
        help=f"写入匹配状态的列名，默认：{DEFAULT_STATUS_COLUMN}。",
    )
    parser.add_argument(
        "--candidate-column",
        default=DEFAULT_CANDIDATE_COLUMN,
        help=f"写入候选结果的列名，默认：{DEFAULT_CANDIDATE_COLUMN}。",
    )
    parser.add_argument(
        "--matched-name-column",
        default=DEFAULT_MATCHED_NAME_COLUMN,
        help=f"写入完整收款开户行的列名，默认：{DEFAULT_MATCHED_NAME_COLUMN}。",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    master_path = args.master.resolve()
    target_path = args.target.resolve()
    if not master_path.is_file():
        print(f"错误：主数据文件不存在：{master_path}", file=sys.stderr)
        return 2
    if not target_path.is_file():
        print(f"错误：待匹配文件不存在：{target_path}", file=sys.stderr)
        return 2
    if master_path == target_path:
        print("错误：主数据文件和待匹配文件不能是同一个文件。", file=sys.stderr)
        return 2

    output_path = args.output or target_path.with_name(f"{target_path.stem}_匹配结果.xlsx")
    output_path = output_path.resolve()
    if output_path in {master_path, target_path}:
        print("错误：结果文件不能覆盖主数据文件或待匹配文件。", file=sys.stderr)
        return 2
    try:
        master_sheets = read_master_data(master_path)
        summary = process_target_file(
            target_path,
            output_path,
            master_sheets,
            args.output_column,
            args.status_column,
            args.candidate_column,
            args.matched_name_column,
        )
    except (OSError, ValueError) as error:
        print(f"处理失败：{error}", file=sys.stderr)
        return 1

    print(f"处理完成：{output_path}")
    print(f"主数据类别：{len(master_sheets)}")
    print(f"处理工作表：{summary['sheets']}")
    print(f"已匹配：{summary.get('matched', 0)}")
    print(f"匹配不唯一：{summary.get('ambiguous', 0)}")
    print(f"未找到：{summary.get('not_found', 0)}")
    print(f"收款开户行为空：{summary.get('empty', 0)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
