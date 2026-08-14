from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from match_bank_codes import (
    _find_candidate_sheets,
    match_bank_name,
    normalize_bank_name,
    process_target_file,
    read_master_data,
)


class BankCodeMatchingTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.directory = Path(self.temporary_directory.name)
        self.master_path = self.directory / "master.xlsx"
        self.target_path = self.directory / "target.xlsx"
        self.output_path = self.directory / "result.xlsx"

        master = Workbook()
        industrial = master.active
        industrial.title = "102中国工商银行"
        industrial.append(["PaySysBnkCode|支付系统行号", "BnkName|银行名称"])
        industrial.append(["102100000001", "中国工商银行股份有限公司北京新华支行"])
        industrial.append(["102100000002", "中国工商银行股份有限公司北京东单支行"])
        industrial.append(["102100000003", "中国工商银行股份有限公司广州天河高新支行"])
        bank_of_china = master.create_sheet("104中国银行")
        bank_of_china.append(["PaySysBnkCode|支付系统行号", "BnkName|银行名称"])
        bank_of_china.append(["104100000001", "中国银行股份有限公司北京通州土桥支行"])
        bank_of_china.append(["104498611366", "中国银行股份有限公司长垣支行"])
        bank_of_china.append(["104498611487", "中国银行股份有限公司长垣蒲东支行"])
        bank_of_china.append(["104000000002", "中国银行甲乙支行"])
        bank_of_china.append(["104000000003", "中国银行甲丙支行"])
        bank_of_china.append(["104000000004", "中国银行乙甲支行"])
        bank_of_china.append(["104000000005", "中国银行股份有限公司广州市环市西路支行"])
        bank_of_china.append(["104000000006", "中国银行股份有限公司广州燕子岗支行"])
        citic = master.create_sheet("302中信银行")
        citic.append(["PaySysBnkCode|支付系统行号", "BnkName|银行名称"])
        citic.append(["302433028718", "中信银行上饶分行"])
        citic.append(["302433228723", "中信银行股份有限公司上饶广信支行"])
        merchants = master.create_sheet("308招商银行")
        merchants.append(["PaySysBnkCode|支付系统行号", "BnkName|银行名称"])
        merchants.append(
            [
                "308100000001",
                "招商银行股份有限公司广州市绿色金融改革创新试验区花都分行",
            ]
        )
        master.save(self.master_path)

        target = Workbook()
        sheet = target.active
        sheet.title = "供应商"
        sheet.append(["供应商", "收款开户行"])
        sheet.append(["A", "工行北京新华支行"])
        sheet.append(["B", "中行北京通州土桥支行"])
        sheet.append(["C", "工行北京支行"])
        sheet.append(["D", "不存在的银行"])
        sheet.append(["E", None])
        sheet.append(["F", "中国银行长垣县支行"])
        target.save(self.target_path)

    def tearDown(self) -> None:
        self.temporary_directory.cleanup()

    def test_abbreviated_bank_names_match_only_safe_codes(self) -> None:
        master_sheets = read_master_data(self.master_path)
        summary = process_target_file(self.target_path, self.output_path, master_sheets)

        self.assertEqual(summary["matched"], 2)
        self.assertEqual(summary["ambiguous"], 2)
        self.assertEqual(summary["not_found"], 1)
        self.assertEqual(summary["empty"], 1)

        result = load_workbook(self.output_path, data_only=True)
        sheet = result["供应商"]
        self.assertEqual(sheet["C2"].value, "102100000001")
        self.assertEqual(sheet["C3"].value, "104100000001")
        self.assertEqual(sheet["D2"].value, "中国工商银行股份有限公司北京新华支行")
        self.assertEqual(sheet["D3"].value, "中国银行股份有限公司北京通州土桥支行")
        self.assertIsNone(sheet["C4"].value)
        self.assertIsNone(sheet["D4"].value)
        self.assertEqual(sheet["E4"].value, "匹配不唯一（2 个候选）")
        self.assertIsNone(sheet["C5"].value)
        self.assertEqual(sheet["E5"].value, "未找到")
        self.assertEqual(sheet["E6"].value, "收款开户行为空")
        self.assertIsNone(sheet["C7"].value)
        self.assertIsNone(sheet["D7"].value)
        self.assertEqual(sheet["E7"].value, "匹配不唯一（2 个候选）")
        self.assertEqual(
            sheet["F7"].value,
            "中国银行股份有限公司长垣支行：104498611366\n"
            "中国银行股份有限公司长垣蒲东支行：104498611487",
        )
        result.close()

    def test_county_suffix_lists_all_candidates_without_discarding_the_county(self) -> None:
        result = match_bank_name("中国银行长垣县支行", read_master_data(self.master_path))

        self.assertIsNone(result.code)
        self.assertEqual(result.status, "匹配不唯一（2 个候选）")
        self.assertEqual(
            result.candidates,
            (
                ("中国银行股份有限公司长垣支行", "104498611366"),
                ("中国银行股份有限公司长垣蒲东支行", "104498611487"),
            ),
        )

    def test_sheet_matching_uses_the_complete_bank_subject_before_abbreviation_fallback(self) -> None:
        master_sheets = read_master_data(self.master_path)

        candidate_sheets = _find_candidate_sheets(
            normalize_bank_name("中国银行长垣县支行"), master_sheets
        )

        self.assertEqual([sheet.title for sheet in candidate_sheets], ["104中国银行"])

    def test_ordered_matching_can_resolve_a_branch_with_an_extra_character(self) -> None:
        result = match_bank_name("中国银行甲无乙支行", read_master_data(self.master_path))

        self.assertEqual(result.code, "104000000002")
        self.assertEqual(result.matched_name, "中国银行甲乙支行")

    def test_directional_matching_preserves_character_order(self) -> None:
        result = match_bank_name("中国银行甲乙无支行", read_master_data(self.master_path))

        self.assertEqual(result.code, "104000000002")
        self.assertEqual(result.matched_name, "中国银行甲乙支行")

    def test_directional_matching_skips_an_unmatched_character(self) -> None:
        result = match_bank_name("工行广州市天河高新支行", read_master_data(self.master_path))

        self.assertEqual(result.code, "102100000003")
        self.assertEqual(result.matched_name, "中国工商银行股份有限公司广州天河高新支行")

    def test_state_machine_skips_an_unmatched_city_character(self) -> None:
        result = match_bank_name("中国银行广州市燕子岗支行", read_master_data(self.master_path))

        self.assertEqual(result.code, "104000000006")
        self.assertEqual(result.matched_name, "中国银行股份有限公司广州燕子岗支行")

    def test_weak_directional_match_is_shown_without_writing_a_code(self) -> None:
        result = match_bank_name("招行广州市天河高新支行", read_master_data(self.master_path))

        self.assertIsNone(result.code)
        self.assertEqual(result.status, "匹配置信度不足（1 个候选）")
        self.assertEqual(
            result.candidates,
            (("招商银行股份有限公司广州市绿色金融改革创新试验区花都分行", "308100000001"),),
        )

    def test_complete_name_matches_before_directional_narrowing(self) -> None:
        result = match_bank_name(
            "中信银行股份有限公司上饶分行", read_master_data(self.master_path)
        )

        self.assertEqual(result.code, "302433028718")
        self.assertEqual(result.matched_name, "中信银行上饶分行")


if __name__ == "__main__":
    unittest.main()
