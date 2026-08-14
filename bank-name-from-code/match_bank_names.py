"""根据支付系统行号补全收款行名称。

这是原有的正向匹配程序。根据收款开户行名称反查支付系统行号，请使用
``match_bank_codes.py``。
"""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from copy import copy
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

MASTER_CODE_HEADERS = ("PaySysBnkCode|支付系统行号", "PaySysBnkCode")
MASTER_NAME_HEADERS = ("BnkName|银行名称", "BnkName")
TARGET_CODE_HEADERS = ("联行号", "支付系统行号", "PaySysBnkCode")
ProgressCallback = Callable[[int, int, dict[str, int]], None]
CancelCallback = Callable[[], bool]
PROJECT_DIRECTORY = Path(__file__).resolve().parent.parent


class MatchingCancelled(Exception):
    """用户在匹配完成前取消了任务。"""


def normalize_code(value: object) -> str:
    """将单元格中的行号统一为 12 位数字文本，保留前导零。"""
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text.zfill(12) if text.isdigit() and len(text) < 12 else text


def header_positions(sheet: Worksheet) -> dict[str, int]:
    """返回首行非空表头及其列号（从 1 开始）。"""
    return {
        str(cell.value).strip(): cell.column
        for row in sheet.iter_rows(min_row=1, max_row=1)
        for cell in row
        if cell.value is not None and str(cell.value).strip()
    }


def find_column(headers: dict[str, int], candidates: tuple[str, ...]) -> int | None:
    """按候选列名顺序返回第一个存在的列号。"""
    for candidate in candidates:
        if candidate in headers:
            return headers[candidate]
    return None


def sheet_prefix(sheet: Worksheet) -> str:
    """工作表名称的前三位数字，例如“001中国人民银行”返回“001”。"""
    title_prefix = sheet.title[:3]
    return title_prefix if title_prefix.isdigit() else ""


def copy_header_style(source: Cell, destination: Cell) -> None:
    """新增列沿用输入列的表头样式。"""
    if source.has_style:
        destination._style = copy(source._style)
    if source.number_format:
        destination.number_format = source.number_format
    if source.font:
        destination.font = copy(source.font)
    if source.fill:
        destination.fill = copy(source.fill)
    if source.border:
        destination.border = copy(source.border)
    if source.alignment:
        destination.alignment = copy(source.alignment)
    if source.protection:
        destination.protection = copy(source.protection)


def read_target_prefixes(
    target_path: Path, target_sheet_names: set[str] | None = None
) -> set[str]:
    """从待匹配文件收集有效支付系统行号的前三位数字。"""
    workbook = load_workbook(target_path, read_only=True, data_only=True)
    prefixes: set[str] = set()
    try:
        for sheet in workbook.worksheets:
            if target_sheet_names is not None and sheet.title not in target_sheet_names:
                continue
            code_column = find_column(header_positions(sheet), TARGET_CODE_HEADERS)
            if code_column is None:
                continue
            for row in sheet.iter_rows(min_row=2, values_only=True):
                code = normalize_code(row[code_column - 1])
                if len(code) == 12 and code.isdigit():
                    prefixes.add(code[:3])
        return prefixes
    finally:
        workbook.close()


def read_master_data(
    master_path: Path, prefixes: set[str] | None = None
) -> tuple[dict[str, str], dict[str, list[str]]]:
    """按行号前三位读取主数据，返回唯一和重复的名称映射。"""
    workbook = load_workbook(master_path, read_only=True, data_only=True)
    try:
        names_by_code: dict[str, set[str]] = defaultdict(set)
        available_prefixes = {
            sheet_prefix(sheet) for sheet in workbook.worksheets if sheet_prefix(sheet)
        }
        for sheet in workbook.worksheets:
            if prefixes is not None and sheet_prefix(sheet) not in prefixes:
                continue
            headers = header_positions(sheet)
            code_column = find_column(headers, MASTER_CODE_HEADERS)
            name_column = find_column(headers, MASTER_NAME_HEADERS)
            if code_column is None or name_column is None:
                continue
            for row in sheet.iter_rows(min_row=2, values_only=True):
                code = normalize_code(row[code_column - 1])
                name_value = row[name_column - 1]
                name = "" if name_value is None else str(name_value).strip()
                if code and name:
                    names_by_code[code].add(name)

        if prefixes:
            missing_prefixes = prefixes - available_prefixes
            if missing_prefixes:
                print(f"提示：主数据中不存在前缀为 {','.join(sorted(missing_prefixes))} 的工作表。")
        elif not names_by_code:
            return {}, {}

        unique_matches: dict[str, str] = {}
        duplicate_matches: dict[str, list[str]] = {}
        for code, names in names_by_code.items():
            ordered_names = sorted(names)
            if len(ordered_names) == 1:
                unique_matches[code] = ordered_names[0]
            else:
                duplicate_matches[code] = ordered_names
        return unique_matches, duplicate_matches
    finally:
        workbook.close()


def _get_or_add_column(sheet: Worksheet, header: str, source_column: int) -> int:
    headers = header_positions(sheet)
    if header in headers:
        return headers[header]
    column = sheet.max_column + 1
    cell = sheet.cell(row=1, column=column, value=header)
    copy_header_style(sheet.cell(row=1, column=source_column), cell)
    return column


def process_target_file(
    target_path: Path,
    output_path: Path,
    unique_matches: dict[str, str],
    duplicate_matches: dict[str, list[str]],
    output_column: str,
    status_column: str,
    *,
    progress_callback: ProgressCallback | None = None,
    cancel_callback: CancelCallback | None = None,
    progress_every: int = 5_000,
    target_sheet_names: set[str] | None = None,
) -> dict[str, int]:
    """保留原有正向匹配的文件写入行为。"""
    if progress_every < 1:
        raise ValueError("progress_every 必须大于 0。")
    workbook = load_workbook(target_path)
    summary: defaultdict[str, int] = defaultdict(int)
    target_sheets: list[tuple[Worksheet, int]] = []
    total_rows = 0
    try:
        if target_sheet_names is not None:
            available = {sheet.title for sheet in workbook.worksheets}
            missing = target_sheet_names - available
            if missing:
                raise ValueError(f"待匹配文件中不存在指定工作表：{'、'.join(sorted(missing))}。")
        for sheet in workbook.worksheets:
            if target_sheet_names is not None and sheet.title not in target_sheet_names:
                continue
            code_column = find_column(header_positions(sheet), TARGET_CODE_HEADERS)
            if code_column is not None:
                target_sheets.append((sheet, code_column))
                total_rows += max(sheet.max_row - 1, 0)
        if not target_sheets:
            raise ValueError(f"待匹配文件中未找到行号列。可识别列名：{'、'.join(TARGET_CODE_HEADERS)}。")
        if progress_callback:
            progress_callback(0, total_rows, dict(summary))

        processed_rows = 0
        for sheet, code_column in target_sheets:
            result_column = _get_or_add_column(sheet, output_column, code_column)
            status_column_index = _get_or_add_column(sheet, status_column, code_column)
            for row_number in range(2, sheet.max_row + 1):
                if cancel_callback and cancel_callback():
                    raise MatchingCancelled("用户已取消匹配。")
                code_cell = sheet.cell(row=row_number, column=code_column)
                code = normalize_code(code_cell.value)
                result_cell = sheet.cell(row=row_number, column=result_column)
                status_cell = sheet.cell(row=row_number, column=status_column_index)
                if not code:
                    result_cell.value = None
                    status_cell.value = "联行号为空"
                    summary["empty"] += 1
                elif code in unique_matches:
                    code_cell.value = code
                    code_cell.number_format = "@"
                    result_cell.value = unique_matches[code]
                    status_cell.value = "已匹配"
                    summary["matched"] += 1
                elif code in duplicate_matches:
                    result_cell.value = "；".join(duplicate_matches[code])
                    status_cell.value = "主数据重复"
                    summary["duplicate"] += 1
                else:
                    result_cell.value = None
                    status_cell.value = "未找到"
                    summary["not_found"] += 1
                processed_rows += 1
                if progress_callback and (
                    processed_rows % progress_every == 0 or processed_rows == total_rows
                ):
                    progress_callback(processed_rows, total_rows, dict(summary))
        workbook.save(output_path)
        summary["sheets"] = len(target_sheets)
        return dict(summary)
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="根据支付系统行号匹配完整收款行名称。")
    parser.add_argument(
        "master", nargs="?", type=Path, default=PROJECT_DIRECTORY / "行名行号2026.03.01.xlsx"
    )
    parser.add_argument(
        "target", nargs="?", type=Path, default=PROJECT_DIRECTORY / "测试文件.xlsx"
    )
    parser.add_argument("-o", "--output", type=Path)
    parser.add_argument("--output-column", default="匹配收款行名称")
    parser.add_argument("--status-column", default="匹配状态")
    args = parser.parse_args()
    master_path = args.master.resolve()
    target_path = args.target.resolve()
    output_path = (args.output or target_path.with_name(f"{target_path.stem}_匹配结果.xlsx")).resolve()
    try:
        prefixes = read_target_prefixes(target_path)
        unique_matches, duplicate_matches = read_master_data(master_path, prefixes)
        summary = process_target_file(
            target_path,
            output_path,
            unique_matches,
            duplicate_matches,
            args.output_column,
            args.status_column,
        )
    except (OSError, ValueError) as error:
        print(f"处理失败：{error}", file=sys.stderr)
        return 1
    print(f"处理完成：{output_path}")
    print(f"已匹配：{summary.get('matched', 0)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
