"""生成百万级联行号匹配压测数据。"""

from __future__ import annotations

import argparse
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill

from match_bank_names import (
    MASTER_CODE_HEADERS,
    MASTER_NAME_HEADERS,
    find_column,
    header_positions,
    normalize_code,
    sheet_prefix,
)


MAX_EXCEL_DATA_ROWS = 1_048_575
PROJECT_DIRECTORY = Path(__file__).resolve().parent.parent


def collect_bank_samples(
    master_path: Path, samples_per_prefix: int = 10
) -> list[tuple[str, str]]:
    """从每个编号工作表提取多条真实行号，并优先覆盖不同末位。"""
    if samples_per_prefix < 1:
        raise ValueError("每个类别的样本数必须大于 0。")

    workbook = load_workbook(master_path, read_only=True, data_only=True)
    samples: list[tuple[str, str]] = []
    sample_codes: set[str] = set()

    for sheet in workbook.worksheets:
        prefix = sheet_prefix(sheet)
        if not prefix:
            continue

        headers = header_positions(sheet)
        code_column = find_column(headers, MASTER_CODE_HEADERS)
        name_column = find_column(headers, MASTER_NAME_HEADERS)
        if code_column is None or name_column is None:
            continue

        samples_by_tail: dict[str, tuple[str, str]] = {}
        extra_samples: list[tuple[str, str]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            code = normalize_code(row[code_column - 1])
            name_value = row[name_column - 1]
            name = "" if name_value is None else str(name_value).strip()
            if not (code.startswith(prefix) and len(code) == 12 and name):
                continue
            if code in sample_codes:
                continue

            sample = (code, name)
            if code[-1] not in samples_by_tail:
                samples_by_tail[code[-1]] = sample
            else:
                extra_samples.append(sample)

            if len(samples_by_tail) == 10 and (
                len(samples_by_tail) + len(extra_samples) >= samples_per_prefix
            ):
                break

        selected = list(samples_by_tail.values())
        selected.extend(extra_samples[: max(0, samples_per_prefix - len(selected))])
        for sample in selected[:samples_per_prefix]:
            sample_codes.add(sample[0])
            samples.append(sample)

    if not samples:
        raise ValueError("未能从主数据中提取有效联行号样本。")

    # 交错不同尾数，避免生成数据中相邻联行号的末位单一。
    samples_by_tail: dict[str, list[tuple[str, str]]] = {str(index): [] for index in range(10)}
    for sample in samples:
        samples_by_tail[sample[0][-1]].append(sample)

    ordered_samples: list[tuple[str, str]] = []
    while any(samples_by_tail.values()):
        for tail in map(str, range(10)):
            if samples_by_tail[tail]:
                ordered_samples.append(samples_by_tail[tail].pop(0))
    return ordered_samples


def incomplete_name(bank_name: str) -> str:
    """生成模拟的不完整收款行名称。"""
    return bank_name[: max(2, min(8, len(bank_name) // 2))]


def make_header_cell(sheet, value: str) -> WriteOnlyCell:
    cell = WriteOnlyCell(sheet, value=value)
    cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell


def available_fallback_path(output_path: Path, row_count: int) -> Path:
    """目标文件被占用时，返回同目录下未被占用的备用文件名。"""
    count_label = f"{row_count // 10_000}万条" if row_count % 10_000 == 0 else f"{row_count}条"
    for index in range(1, 10_000):
        suffix = "" if index == 1 else f"_{index}"
        candidate = output_path.with_name(f"{output_path.stem}_{count_label}{suffix}.xlsx")
        if not candidate.exists():
            return candidate
    raise OSError("无法找到可用的备用结果文件名。")


def generate_test_file(master_path: Path, output_path: Path, row_count: int) -> tuple[int, Path]:
    if not 1 <= row_count <= MAX_EXCEL_DATA_ROWS:
        raise ValueError(f"数据行数必须在 1 到 {MAX_EXCEL_DATA_ROWS:,} 之间。")
    if master_path.resolve() == output_path.resolve():
        raise ValueError("主数据文件和测试文件不能相同。")
    if not master_path.is_file():
        raise FileNotFoundError(f"主数据文件不存在：{master_path}")

    samples = collect_bank_samples(master_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(f"{output_path.stem}.__generating__.xlsx")
    if temporary_path.exists():
        temporary_path.unlink()

    try:
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet("Sheet1")
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 18
        sheet.column_dimensions["C"].width = 18
        sheet.append(
            [
                make_header_cell(sheet, "客户名称"),
                make_header_cell(sheet, "原收款行名称"),
                make_header_cell(sheet, "联行号"),
            ]
        )

        for index in range(1, row_count + 1):
            code, bank_name = samples[(index - 1) % len(samples)]
            sheet.append(
                [
                    f"压测客户{index:07d}",
                    incomplete_name(bank_name),
                    code,
                ]
            )
            if index % 100_000 == 0 or index == row_count:
                print(f"已生成 {index:,}/{row_count:,} 条")

        sheet.auto_filter.ref = f"A1:C{row_count + 1}"
        workbook.save(temporary_path)
        try:
            os.replace(temporary_path, output_path)
            final_path = output_path
        except PermissionError:
            final_path = available_fallback_path(output_path, row_count)
            os.replace(temporary_path, final_path)
            print(f"提示：{output_path.name} 正被占用，已另存为：{final_path.name}")
    except Exception:
        if temporary_path.exists():
            temporary_path.unlink()
        raise

    print(f"已使用 {len(samples)} 个主数据样本。")
    return len(samples), final_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="生成用于联行号匹配的百万级测试 Excel 文件。")
    parser.add_argument(
        "--master",
        type=Path,
        default=PROJECT_DIRECTORY / "行名行号2026.03.01.xlsx",
        help="银行行名行号主数据文件。",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=PROJECT_DIRECTORY / "测试文件.xlsx",
        help="生成的测试文件路径。",
    )
    parser.add_argument(
        "--rows",
        type=int,
        default=10_000,
        help="生成的数据行数，默认：10,000。",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    master_path = args.master.resolve()
    output_path = args.output.resolve()

    try:
        sample_count, final_path = generate_test_file(master_path, output_path, args.rows)
    except (OSError, ValueError) as error:
        print(f"生成失败：{error}")
        return 1

    print(f"生成完成：{final_path}")
    print(f"主数据样本数：{sample_count}")
    print(f"数据行数：{args.rows:,}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
